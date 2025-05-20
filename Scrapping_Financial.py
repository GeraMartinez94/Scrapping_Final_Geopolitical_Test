import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

<<<<<<< HEAD
# --- Configuración ---
=======
# -- Configuración -- #
>>>>>>> 95a8f6ce99b35a733e412b5d0c85f07675935a56
URL = "https://www.geopriskindex.com/results-final-risk-index/"
EXCEL_FILENAME = "formato_wide_geopriskinder.xlsx"
SHEET_NAME = 'Índice de Riesgo Global'
COUNTRY_COLUMN_NAME = '\ufeffCountry'

<<<<<<< HEAD
=======
try:
    #-- Obtiene los datos de la web --#
    response = requests.get(URL)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')
    table = soup.find('table')
>>>>>>> 95a8f6ce99b35a733e412b5d0c85f07675935a56

#######################################################
  #Obtiene el objeto tabla BeautifulSoup de la URL.#
#######################################################

def extraer_tabla(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        return soup.find('table')
    except requests.exceptions.RequestException as e:
        print(f"Error al acceder a la página: {e}")
        return None

##################################################################
 # Extrae los encabezados y los datos de las filas de la tabla. #
##################################################################

def extraer_datos_tabla(table):
    headers = []
    data = []
    if table:
        headers = [th.text.strip() for th in table.find_all('th')]
<<<<<<< HEAD
=======

        #-- Extraer datos de filas --#
        data = []
>>>>>>> 95a8f6ce99b35a733e412b5d0c85f07675935a56
        tbody = table.find('tbody')
        if tbody:
            for row in tbody.find_all('tr')[1:]:
                cols = [td.text.strip() for td in row.find_all('td')]
                data.append(cols)
        else:
            print("No se encontró el cuerpo de la tabla (tbody).")
    return headers, data

<<<<<<< HEAD
##########################################
    #Crea un DataFrame en formato ancho.#
##########################################

def crear_dataframe_ancho(df, country_col_name):
    if df is None or df.empty:
        return pd.DataFrame()
    unique_years = df['Year'].unique()
    unique_countries = df[country_col_name].unique()
    df_wide = pd.DataFrame()
    df_wide['Date'] = [f"{year}-mm-dd" for year in unique_years for _ in unique_countries]
    df_wide['Country'] = [country for _ in unique_years for country in unique_countries]
=======
        #-- Crea DataFrame inicial --#
        df = pd.DataFrame(data, columns=headers)
        print("Encabezados del DataFrame:", headers)

        #-- Crear DataFrame en formato ancho --#
        unique_years = df['Year'].unique()
        unique_countries = df[COUNTRY_COLUMN_NAME].unique()
        df_wide = pd.DataFrame()
        df_wide['Date'] = [f"{year}-mm-dd" for year in unique_years for _ in unique_countries]
        df_wide['Country'] = [country for _ in unique_years for country in unique_countries]
>>>>>>> 95a8f6ce99b35a733e412b5d0c85f07675935a56

    for header in df.columns:
        if header not in [country_col_name, 'Year', 'Region']:
            values = []
            for year in unique_years:
                df_year = df[df['Year'] == year]
                country_value_map = df_year.set_index(country_col_name)[header].to_dict()
                for country in unique_countries:
                    values.append(country_value_map.get(country, ''))
            df_wide[header] = values
    return df_wide

<<<<<<< HEAD
####################################################################
    #Guarda el DataFrame en Excel y ajusta el ancho de las columnas.#
####################################################################
  
def guardar_y_ajustar_excel(df_wide, filename, sheet_name):
    if df_wide is None or df_wide.empty:
        print("No hay datos para guardar en Excel.")
        return
    df_wide.to_excel(filename, index=False, sheet_name=sheet_name)
    print(f"Se creo el archivo Excel: {filename}")
=======
        #-- Guarda el DataFrame en  archivo Excel --#
        df_wide.to_excel(EXCEL_FILENAME, index=False, sheet_name=SHEET_NAME)
        print(f"Se creo el archivo Excel: {EXCEL_FILENAME}")

        #-- Ajusta ancho de columnas en Excel --#
        workbook = load_workbook(EXCEL_FILENAME)
        sheet = workbook[SHEET_NAME]
>>>>>>> 95a8f6ce99b35a733e412b5d0c85f07675935a56

    try:
        workbook = load_workbook(filename)
        sheet = workbook[sheet_name]
        for column_cells in sheet.columns:
            max_len = 0
            column = [cell.value for cell in column_cells if cell.value is not None]
            if column:
                max_len = max(len(str(value)) for value in column)
            adjusted_width = max_len + 2
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
        workbook.save(filename)
        print(f"Se ajustó el ancho de las columnas en: {filename}")
    except Exception as e:
        print(f"Error al ajustar el ancho de las columnas: {e}")

if __name__ == "__main__":
    table = extraer_tabla(URL)
    if table:
        headers, data = extraer_datos_tabla(table)
        if headers and data:
            df = pd.DataFrame(data, columns=headers)
            print("Encabezados del DataFrame:", headers)
            df_wide = crear_dataframe_ancho(df, COUNTRY_COLUMN_NAME)
            guardar_y_ajustar_excel(df_wide, EXCEL_FILENAME, SHEET_NAME)
        else:
            print("No se pudieron extraer encabezados o datos de la tabla.")
    else:
<<<<<<< HEAD
        print("No se encontró ninguna tabla en la página.")
=======
        print("No se encontró ninguna tabla en la página.")

except requests.exceptions.RequestException as e:
    print(f"Error al acceder a la página: {e}")
except Exception as e:
    print(f"Ocurrió un error: {e}")
>>>>>>> 95a8f6ce99b35a733e412b5d0c85f07675935a56
